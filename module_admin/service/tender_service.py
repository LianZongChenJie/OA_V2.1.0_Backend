import os
import logging
from typing import Any, List
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import pandas as pd
from io import BytesIO
from fastapi import UploadFile
from openpyxl.styles import PatternFill, Font

from utils.file_util import allowed_file, generate_file_path, generate_file_path_without_id, save_upload_file, ALLOWED_EXTENSIONS, get_file_ext
from common.vo import PageModel, CrudResponseModel
from module_admin.dao.tender_dao import TenderDao
from module_admin.entity.do.user_do import SysUser
from module_admin.entity.vo.tender_vo import (
    TenderPageQueryModel, AddTenderModel, EditTenderModel, DeleteTenderModel,
    AddTenderAttachmentModel, DeleteTenderAttachmentModel, TenderImportTempModel,
    TenderImportResponseModel
)
from exceptions.exception import ServiceException

# 初始化日志器
logger = logging.getLogger(__name__)

class TenderService:
    """招投标管理服务层（最终适配版）"""

    @classmethod
    async def upload_tender_attachment_services(
            cls, file: UploadFile
    ) -> CrudResponseModel:
        """
        上传投标附件（仅上传文件，不关联数据库）
        文件关联逻辑由前端在新增投标信息接口中处理
        :param file: 上传的文件对象
        :return: 上传结果
        """
        # 1. 校验文件格式
        logger.info(f'校验文件格式：{file.filename}')
        if not allowed_file(file.filename):
            raise ServiceException(
                message=f'文件格式不支持，仅支持：{", ".join(ALLOWED_EXTENSIONS)}，当前文件：{file.filename}'
            )

        absolute_path = None
        try:
            # 2. 生成文件存储路径（使用临时目录，不依赖投标ID）
            logger.info(f'生成文件存储路径：{file.filename}')
            relative_path, absolute_path = generate_file_path_without_id(file.filename)
            logger.info(f'文件存储路径：相对路径={relative_path}，绝对路径={absolute_path}')

            # 确保目录存在
            file_dir = os.path.dirname(absolute_path)
            if not os.path.exists(file_dir):
                os.makedirs(file_dir, exist_ok=True)
                logger.info(f'创建目录：{file_dir}')

            # 3. 保存文件到本地
            logger.info(f'保存文件到本地：{absolute_path}')
            file_size = await save_upload_file(file, absolute_path)
            logger.info(f'文件保存成功，大小：{file_size} 字节')

            # 返回文件信息，供前端后续关联使用
            return CrudResponseModel(
                is_success=True,
                message='附件上传成功',
                result={
                    'file_name': file.filename,
                    'file_path': relative_path,
                    'file_size': file_size,
                    'file_ext': get_file_ext(file.filename) or '',
                    'file_mime': file.content_type or 'application/octet-stream'
                }
            )
        except Exception as e:
            logger.error(f'附件上传失败：{str(e)}', exc_info=True)

            # 上传失败时删除已保存的文件
            if absolute_path and os.path.exists(absolute_path):
                os.remove(absolute_path)
                logger.info(f'删除已保存的文件：{absolute_path}')

            raise ServiceException(message=f'附件上传失败：{str(e)}') from e

    @classmethod
    async def get_tender_list_services(
            cls, query_db: AsyncSession, query_object: TenderPageQueryModel, is_page: bool = False
    ) -> PageModel | list[dict[str, Any]]:
        """获取投标列表"""
        try:
            result = await TenderDao.get_tender_list(query_db, query_object, is_page)
            return result
        except Exception as e:
            logger.error(f'获取投标列表失败：{str(e)}', exc_info=True)
            raise ServiceException(message=f'获取投标列表失败：{str(e)}') from e

    @classmethod
    async def tender_detail_services(cls, query_db: AsyncSession, tender_id: int) -> dict[str, Any]:
        """获取投标详情"""
        tender_info = await TenderDao.get_tender_detail_by_id(query_db, tender_id)
        if not tender_info:
            raise ServiceException(message=f'投标信息不存在，ID：{tender_id}')

        try:
            attachments = await TenderDao.get_tender_attachments(query_db, tender_id)
            return {'tender': tender_info, 'attachments': attachments}
        except Exception as e:
            raise ServiceException(message=f'获取投标附件失败：{str(e)}') from e

    @classmethod
    async def add_tender_services(cls, query_db: AsyncSession, page_object: AddTenderModel) -> CrudResponseModel:
        """新增投标（含附件）"""
        # 项目名称验重
        if page_object.project_name:
            existing_tender = await TenderDao.get_tender_by_project_name(query_db, page_object.project_name)
            if existing_tender:
                raise ServiceException(message=f'项目名称已存在：{page_object.project_name}')

        # 日期转换
        if page_object.purchase_date:
            try:
                dt = datetime.strptime(page_object.purchase_date, '%Y-%m-%d')
                page_object.purchase_date = dt.date().strftime('%Y-%m-%d')  # 转为字符串匹配VO模型
            except ValueError:
                raise ServiceException(message='purchase_date 日期格式错误，仅支持 YYYY-MM-DD')

        if page_object.bid_opening_date:
            try:
                dt = datetime.strptime(page_object.bid_opening_date, '%Y-%m-%d')
                page_object.bid_opening_date = dt.date().strftime('%Y-%m-%d')  # 转为字符串匹配VO模型
            except ValueError:
                raise ServiceException(message='bid_opening_date 日期格式错误，仅支持 YYYY-MM-DD')

        # 枚举字段转换（支持 '是', 'YES', 'yes', 'Y', '1', 'true', 'TRUE' 等多种格式）
        tender_dict = page_object.model_dump()
        yes_values = ['是', 'YES', 'yes', 'Y', '1', 'true', 'TRUE']
        enum_mapping = {
            'is_tender_submitted': '是' if tender_dict.get('is_tender_submitted') in yes_values else '否',
            'has_tender_invoice': '是' if tender_dict.get('has_tender_invoice') in yes_values else '否',
            'is_deposit_paid': '是' if tender_dict.get('is_deposit_paid') in yes_values else '否',
            'is_deposit_refunded': '是' if tender_dict.get('is_deposit_refunded') in yes_values else '否'
        }
        for field, value in enum_mapping.items():
            setattr(page_object, field, value)

        # 处理空字符串的日期时间字段，转换为 None
        if page_object.deposit_paid_time == '':
            page_object.deposit_paid_time = None
        
        # 时间字段处理（转为字符串匹配VO模型）
        page_object.create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        page_object.update_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        page_object.delete_time = 0

        try:
            # 保存投标基本信息，获取返回的投标对象
            tender_result = await TenderDao.add_tender_dao(query_db, page_object)
            tender_id = tender_result.id
            
            # 如果有附件，批量保存附件
            if page_object.attachments and len(page_object.attachments) > 0:
                for attachment in page_object.attachments:
                    # 构建附件新增模型
                    add_attachment_model = AddTenderAttachmentModel(
                        project_tender_id=tender_id,
                        file_name=attachment.file_name or '',
                        file_path=attachment.file_path or '',
                        file_size=attachment.file_size or 0,
                        file_ext=attachment.file_ext or '',
                        file_mime=attachment.file_mime or '',
                        sort=attachment.sort or 0,
                        delete_time=0
                    )
                    await TenderDao.add_tender_attachment_dao(query_db, add_attachment_model)
            
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='新增成功', result={'tender_id': tender_id})
        except Exception as e:
            await query_db.rollback()
            raise ServiceException(message=f'新增失败：{str(e)}') from e

    @classmethod
    async def edit_tender_services(cls, query_db: AsyncSession, page_object: EditTenderModel) -> CrudResponseModel:
        """编辑投标"""
        tender_info = await TenderDao.get_tender_detail_by_id(query_db, page_object.id)
        if not tender_info:
            raise ServiceException(message=f'投标信息不存在，ID：{page_object.id}')

        # 项目名称验重（排除当前记录）
        if page_object.project_name:
            existing_tender = await TenderDao.get_tender_by_project_name(query_db, page_object.project_name, exclude_id=page_object.id)
            if existing_tender:
                raise ServiceException(message=f'项目名称已存在：{page_object.project_name}')

        # 日期格式转换
        if page_object.purchase_date:
            try:
                dt = datetime.strptime(page_object.purchase_date, '%Y-%m-%d')
                page_object.purchase_date = dt.date().strftime('%Y-%m-%d')
            except ValueError:
                raise ServiceException(message='purchase_date 日期格式错误，仅支持 YYYY-MM-DD')

        if page_object.bid_opening_date:
            try:
                dt = datetime.strptime(page_object.bid_opening_date, '%Y-%m-%d')
                page_object.bid_opening_date = dt.date().strftime('%Y-%m-%d')
            except ValueError:
                raise ServiceException(message='bid_opening_date 日期格式错误，仅支持 YYYY-MM-DD')

        # 枚举字段转换（支持 '是', 'YES', 'yes', 'Y', '1', 'true', 'TRUE' 等多种格式）
        tender_dict = page_object.model_dump()
        yes_values = ['是', 'YES', 'yes', 'Y', '1', 'true', 'TRUE']
        enum_mapping = {
            'is_tender_submitted': '是' if tender_dict.get('is_tender_submitted') in yes_values else '否',
            'has_tender_invoice': '是' if tender_dict.get('has_tender_invoice') in yes_values else '否',
            'is_deposit_paid': '是' if tender_dict.get('is_deposit_paid') in yes_values else '否',
            'is_deposit_refunded': '是' if tender_dict.get('is_deposit_refunded') in yes_values else '否'
        }
        for field, value in enum_mapping.items():
            setattr(page_object, field, value)

        # 处理空字符串的日期时间字段，转换为 None
        if page_object.deposit_paid_time == '':
            page_object.deposit_paid_time = None
        
        if page_object.create_time == '':
            page_object.create_time = None

        # 更新时间
        page_object.update_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        try:
            # 更新投标基本信息
            await TenderDao.edit_tender_dao(query_db, page_object)
            
            # 如果有附件，先删除原有附件（软删除），再新增新附件
            if page_object.attachments and len(page_object.attachments) > 0:
                # 获取现有附件列表
                existing_attachments = await TenderDao.get_tender_attachments(query_db, page_object.id)
                if existing_attachments:
                    # 软删除现有附件
                    attachment_ids = [att.id for att in existing_attachments]
                    delete_time = int(datetime.now().timestamp())
                    await TenderDao.batch_delete_tender_attachments_dao(query_db, attachment_ids, delete_time)
                
                # 新增新附件
                for attachment in page_object.attachments:
                    add_attachment_model = AddTenderAttachmentModel(
                        project_tender_id=page_object.id,
                        file_name=attachment.file_name or '',
                        file_path=attachment.file_path or '',
                        file_size=attachment.file_size or 0,
                        file_ext=attachment.file_ext or '',
                        file_mime=attachment.file_mime or '',
                        sort=attachment.sort or 0,
                        delete_time=0
                    )
                    await TenderDao.add_tender_attachment_dao(query_db, add_attachment_model)
            
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='修改成功')
        except Exception as e:
            await query_db.rollback()
            raise ServiceException(message=f'修改失败：{str(e)}') from e

    @classmethod
    async def delete_tender_services(
            cls, query_db: AsyncSession, page_object: DeleteTenderModel
    ) -> CrudResponseModel:
        """删除投标（软删除）"""
        if not page_object.ids:
            raise ServiceException(message='传入投标 id 为空')

        try:
            tender_ids = [int(id_str.strip()) for id_str in page_object.ids.split(',') if id_str.strip()]
            if not tender_ids:
                raise ServiceException(message='投标ID格式错误，应为数字，多个用逗号分隔')

            delete_time = int(datetime.now().timestamp())
            await TenderDao.delete_tender_dao(query_db, tender_ids, delete_time)
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='删除成功')
        except ValueError:
            raise ServiceException(message='投标ID必须为数字，多个用逗号分隔')
        except Exception as e:
            await query_db.rollback()
            raise ServiceException(message=f'删除失败：{str(e)}') from e

    @classmethod
    async def add_tender_attachment_services(
            cls, query_db: AsyncSession, page_object: AddTenderAttachmentModel
    ) -> CrudResponseModel:
        """新增投标附件"""
        tender_info = await TenderDao.get_tender_detail_by_id(query_db, page_object.project_tender_id)
        if not tender_info:
            raise ServiceException(message=f'关联的投标信息不存在，ID：{page_object.project_tender_id}')

        try:
            await TenderDao.add_tender_attachment_dao(query_db, page_object)
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='新增附件成功')
        except Exception as e:
            await query_db.rollback()
            logger.error(f'新增附件失败：{str(e)}', exc_info=True)
            raise ServiceException(message=f'新增附件失败：{str(e)}') from e

    @classmethod
    async def delete_tender_attachment_services(
            cls, query_db: AsyncSession, page_object: DeleteTenderAttachmentModel
    ) -> CrudResponseModel:
        """删除投标附件（软删除）"""
        if not page_object.ids:
            raise ServiceException(message='传入附件 id 为空')

        try:
            await TenderDao.delete_tender_attachment_dao(query_db, page_object)
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='删除附件成功')
        except Exception as e:
            await query_db.rollback()
            raise ServiceException(message=f'删除附件失败：{str(e)}') from e

    @classmethod
    async def generate_tender_import_template(cls) -> BytesIO:
        """生成投标导入Excel模板"""
        try:
            # 定义字段映射
            field_mapping = {
                'month': '月份(必填，格式：YYYY-MM)',
                'customer_name': '客户名称(必填)',
                'project_name': '项目名称(必填)',
                'tender_leader_id': '标书负责人ID',
                'tender_leader': '投标负责人',
                'purchase_date': '购买日期(格式：YYYY-MM-DD)',
                'tender_agency': '招标机构',
                'project_cycle': '项目周期',
                'project_cycle_num': '项目周期(月)',
                'shortlisted_countries': '入围家数(数字)',
                'budget_amount': '预算金额(数字，元)',
                'bid_opening_date': '开标日期(格式：YYYY-MM-DD)',
                'is_tender_submitted': '是否投标(必填，是/否)',
                'non_tender_reason': '未投原因',
                'tender_document_fee': '标书款(数字，元)',
                'has_tender_invoice': '标书款发票(是/否)',
                'is_deposit_paid': '是否缴纳保证金(是/否)',
                'tender_deposit': '投标保证金(数字，元)',
                'deposit_account_name': '保证金账户名',
                'deposit_bank': '保证金开户行',
                'deposit_account_no': '保证金账号',
                'is_deposit_refunded': '是否退回保证金(是/否)',
                'bid_result': '中标结果(待开标/中标/未中标/流标/未投标)',
                'bid_service_fee': '中标服务费(数字，元)',
                'remark': '备注',
                'sort': '排序(数字)'
            }

            headers = list(field_mapping.keys())
            header_notes = list(field_mapping.values())

            # 示例数据
            sample_data = [
                {
                    'month': '2024-01',
                    'customer_name': '北京XX科技有限公司',
                    'project_name': '智慧园区建设项目',
                    'tender_leader': '张三',
                    'purchase_date': '2024-01-05',
                    'tender_agency': 'XX招标代理有限公司',
                    'project_cycle': '12个月',
                    'shortlisted_countries': '5',
                    'budget_amount': '1000000',
                    'bid_opening_date': '2024-01-20',
                    'is_tender_submitted': '是',
                    'non_tender_reason': '',
                    'tender_document_fee': '800',
                    'has_tender_invoice': '是',
                    'is_deposit_paid': '是',
                    'tender_deposit': '50000',
                    'deposit_account_name': '北京XX科技有限公司',
                    'deposit_bank': '中国建设银行北京海淀支行',
                    'deposit_account_no': '6222081100001234567',
                    'is_deposit_refunded': '是',
                    'bid_result': '中标',
                    'bid_service_fee': '5000',
                    'remark': '这是一个备注示例',
                    'sort': '1'
                }
            ]

            # 构造DataFrame
            df_notes = pd.DataFrame([header_notes], columns=headers)
            df_data = pd.DataFrame(sample_data, columns=headers)
            df = pd.concat([df_notes, df_data], ignore_index=True)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='投标信息模板', index=False, header=True)
                worksheet = writer.sheets['投标信息模板']

                # 样式设置
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                for cell in worksheet[2]:
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

                # 列宽自适应
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

                # 设置日期列的单元格格式为文本格式，防止Excel自动转换日期格式
                date_columns = ['month', 'purchase_date', 'bid_opening_date']
                for col_name in date_columns:
                    if col_name in field_mapping:
                        col_index = headers.index(col_name)
                        col_letter = worksheet.cell(row=1, column=col_index + 1).column_letter
                        # 设置整列（包括空行）格式为文本
                        # 先填充100行空数据来确保列格式生效
                        for row_num in range(3, 502):  # 填充第2行到第101行
                            cell = worksheet.cell(row=row_num, column=col_index + 1)
                            cell.number_format = '@'  # 文本格式
                            if row_num == 3:  # 保留第一行样例数据
                                cell.value = sample_data[0][col_name]
                            else:
                                cell.value = ''

            output.seek(0)
            return output
        except Exception as e:
            logger.error(f"生成模板失败: {str(e)}", exc_info=True)
            raise ServiceException(message=f'生成模板失败：{str(e)}') from e

    @classmethod
    async def import_tender_services(
            cls, query_db: AsyncSession, file: UploadFile
    ) -> TenderImportResponseModel:
        """批量导入投标信息"""
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise ServiceException(message='仅支持Excel文件（.xlsx/.xls）导入')

        try:
            contents = await file.read()
            df = pd.read_excel(
                BytesIO(contents),
                sheet_name=0,
                header=0,
                skiprows=[1],
                dtype=str
            )

            # 数据清洗
            df.columns = df.columns.str.strip()
            df = df.dropna(how='all').fillna('').reset_index(drop=True)

            # 字段对齐
            model_fields = TenderImportTempModel.model_fields.keys()
            for field in model_fields:
                if field not in df.columns:
                    df[field] = ''
            df = df[model_fields]
            excel_data = df.to_dict('records')

            success_count = 0
            fail_count = 0
            fail_data = []

            for idx, row in enumerate(excel_data):
                row_num = idx + 3
                try:
                    # 必填字段校验
                    required_fields = ['month', 'customer_name', 'project_name', 'is_tender_submitted']
                    missing_fields = [f for f in required_fields if not str(row.get(f, '')).strip()]
                    if missing_fields:
                        raise ServiceException(f'必填字段为空：{",".join(missing_fields)}')

                    # 模型校验
                    import_model = TenderImportTempModel(**row)

                    # 初始化新增模型
                    add_model = AddTenderModel()

                    # 日期转换（转为字符串匹配VO模型）
                    if import_model.purchase_date.strip():
                        try:
                            dt = datetime.strptime(import_model.purchase_date.strip(), '%Y-%m-%d')
                            add_model.purchase_date = dt.date().strftime('%Y-%m-%d')
                        except ValueError:
                            raise ServiceException(f'购买日期格式错误，应为YYYY-MM-DD')

                    if import_model.bid_opening_date.strip():
                        try:
                            dt = datetime.strptime(import_model.bid_opening_date.strip(), '%Y-%m-%d')
                            add_model.bid_opening_date = dt.date().strftime('%Y-%m-%d')
                        except ValueError:
                            raise ServiceException(f'开标日期格式错误，应为YYYY-MM-DD')

                    # 枚举转换
                    enum_fields = ['is_tender_submitted', 'has_tender_invoice', 'is_deposit_paid', 'is_deposit_refunded']
                    for field in enum_fields:
                        value = getattr(import_model, field, '').strip()
                        setattr(add_model, field, '是' if value in ['是', 'YES', 'yes'] else '否')

                    # 数字转换
                    add_model.shortlisted_countries = import_model.shortlisted_countries.strip() or None
                    add_model.budget_amount = float(import_model.budget_amount.strip()) if import_model.budget_amount.strip() else None
                    add_model.tender_document_fee = float(import_model.tender_document_fee.strip()) if import_model.tender_document_fee.strip() else None
                    add_model.tender_deposit = float(import_model.tender_deposit.strip()) if import_model.tender_deposit.strip() else None
                    add_model.bid_service_fee = float(import_model.bid_service_fee.strip()) if import_model.bid_service_fee.strip() else None
                    add_model.sort = int(import_model.sort.strip()) if import_model.sort.strip() else 0

                    # 普通字段赋值
                    add_model.month = import_model.month.strip()
                    add_model.customer_name = import_model.customer_name.strip()
                    add_model.project_name = import_model.project_name.strip()
                    
                    # 根据 tender_leader 姓名查询 user_id
                    tender_leader_name = import_model.tender_leader.strip() if hasattr(import_model, 'tender_leader') and import_model.tender_leader.strip() else None
                    if tender_leader_name:
                        try:
                            user_query = select(SysUser.user_id).where(
                                SysUser.nick_name == tender_leader_name,
                                SysUser.del_flag == '0'
                            )
                            user_result = await query_db.execute(user_query)
                            user_id = user_result.scalar_one_or_none()
                            if user_id:
                                add_model.tender_leader_id = str(user_id)
                            else:
                                logger.warning(f"未找到负责人 '{tender_leader_name}' 对应的用户ID")
                        except Exception as e:
                            logger.error(f"查询负责人 '{tender_leader_name}' 的 user_id 失败: {str(e)}")
                    
                    add_model.tender_leader = tender_leader_name or None
                    add_model.tender_agency = import_model.tender_agency.strip() or None
                    add_model.project_cycle = import_model.project_cycle.strip() or None
                    add_model.project_cycle_num = int(import_model.project_cycle_num.strip()) if import_model.project_cycle_num.strip() else None
                    add_model.non_tender_reason = import_model.non_tender_reason.strip() or None
                    add_model.deposit_account_name = import_model.deposit_account_name.strip() or None
                    add_model.deposit_bank = import_model.deposit_bank.strip() or None
                    add_model.deposit_account_no = import_model.deposit_account_no.strip() or None
                    add_model.bid_result = import_model.bid_result.strip() or None
                    add_model.remark = import_model.remark.strip() or None

                    # 时间字段（转为字符串匹配VO模型）
                    add_model.create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    add_model.update_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    add_model.delete_time = 0

                    # 插入数据库
                    await TenderDao.add_tender_dao(query_db, add_model)
                    success_count += 1

                except Exception as e:
                    fail_count += 1
                    fail_data.append({
                        'row': row_num,
                        'data': row,
                        'reason': str(e)
                    })
                    logger.error(f"第{row_num}行导入失败: {str(e)}")
                    continue

            await query_db.commit()
            logger.info(f"导入完成：成功{success_count}条，失败{fail_count}条")

            return TenderImportResponseModel(
                is_success=True,
                message=f'导入完成：成功{success_count}条，失败{fail_count}条',
                success_count=success_count,
                fail_count=fail_count,
                fail_data=fail_data
            )

        except Exception as e:
            await query_db.rollback()
            logger.error(f"导入投标信息失败: {str(e)}", exc_info=True)
            raise ServiceException(message=f'导入失败：{str(e)}') from e